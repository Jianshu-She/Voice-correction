"""
Voice Correction Pipeline v2.0 (WavLM Fine-tuned)
==================================================
Phoneme-level English pronunciation assessment using a fine-tuned WavLM-Large
backbone + MLP scoring head, trained on 11K children's speech samples.

Improvements over v1.0:
- Fine-tuned WavLM-Large backbone (vs frozen wav2vec2 + GOP threshold)
- Learned phoneme scoring (vs heuristic GOP threshold)
- AUC 0.870 (vs 0.738), F1 0.595 (vs 0.476), Pearson 0.645 (vs 0.372)

Usage:
    # Single file
    python pipeline_v2.py --audio audio.mp3 --text "Hello, Peter."

    # Batch mode
    python pipeline_v2.py --batch --input eval_log.xlsx --audio-dir audio_files/ --output results.json

    # Evaluate against ground truth
    python pipeline_v2.py --batch --input eval_log.xlsx --audio-dir audio_files/ --evaluate
"""

import argparse
import json
import re
import sys
import warnings
from pathlib import Path

import numpy as np
import torch
import torch.nn as nn
import torchaudio
import torchaudio.functional as F_audio
from g2p_en import G2p
from huggingface_hub import hf_hub_download
from transformers import Wav2Vec2FeatureExtractor, Wav2Vec2ForCTC, WavLMModel

warnings.filterwarnings("ignore")

# ============================================================
# Constants
# ============================================================
SAMPLE_RATE = 16000
DEFAULT_PHERR_THRESHOLD = 0.70  # Calibrated on validation set

ARPABET_TO_IPA = {
    "aa": ["ɑː", "ɑ", "ɒ", "a"], "ae": ["æ"], "ah": ["ʌ", "ə", "ɐ"],
    "ao": ["ɔː", "ɔ", "ɒ"], "aw": ["aʊ"], "ax": ["ə", "ɐ", "ʌ"], "ay": ["aɪ"],
    "b": ["b"], "ch": ["tʃ"], "d": ["d"], "dh": ["ð"],
    "eh": ["ɛ", "e"], "er": ["ɜː", "ɝ", "ɚ", "ɜ"], "ey": ["eɪ"],
    "f": ["f"], "g": ["ɡ", "g"], "hh": ["h"],
    "ih": ["ɪ", "ᵻ"], "iy": ["iː", "i"],
    "ir": ["ɪɹ"], "jh": ["dʒ"], "k": ["k"], "l": ["l"],
    "m": ["m"], "n": ["n"], "ng": ["ŋ"],
    "ow": ["oʊ", "o", "əʊ"], "oy": ["ɔɪ"],
    "p": ["p"], "r": ["ɹ", "r"], "s": ["s"], "sh": ["ʃ"],
    "t": ["t"], "th": ["θ"], "uh": ["ʊ"], "uw": ["uː", "u"],
    "ur": ["ʊɹ"], "v": ["v"], "w": ["w"], "y": ["j"],
    "z": ["z"], "zh": ["ʒ"],
    "ar": ["ɑːɹ"], "oo": ["ʊ", "uː"], "dr": ["dɹ"], "tr": ["tɹ"],
    "ts": ["ts"], "dz": ["dz"],
}

ALL_PHONES = sorted(ARPABET_TO_IPA.keys())
PHONE_TO_ID = {ph: i for i, ph in enumerate(ALL_PHONES)}
N_PHONE_TYPES = len(ALL_PHONES)


# ============================================================
# MLP Scoring Head (must match training architecture)
# ============================================================
class PhoneScorerHead(nn.Module):
    def __init__(self, hidden_dim=1024, n_phone_types=N_PHONE_TYPES,
                 phone_emb_dim=32, mlp_dim=512):
        super().__init__()
        self.phone_emb = nn.Embedding(n_phone_types, phone_emb_dim)
        input_dim = hidden_dim + phone_emb_dim + 2  # +2 for GOP and n_frames

        self.shared = nn.Sequential(
            nn.Linear(input_dim, mlp_dim),
            nn.BatchNorm1d(mlp_dim),
            nn.GELU(),
            nn.Dropout(0.3),
            nn.Linear(mlp_dim, mlp_dim),
            nn.BatchNorm1d(mlp_dim),
            nn.GELU(),
            nn.Dropout(0.3),
            nn.Linear(mlp_dim, 256),
            nn.BatchNorm1d(256),
            nn.GELU(),
            nn.Dropout(0.2),
        )
        self.score_head = nn.Sequential(nn.Linear(256, 64), nn.GELU(), nn.Linear(64, 1))
        self.pherr_head = nn.Sequential(nn.Linear(256, 64), nn.GELU(), nn.Linear(64, 1))

    def forward(self, h, phone_id, gop, n_frames):
        emb = self.phone_emb(phone_id)
        x = torch.cat([h, emb, gop.unsqueeze(-1), n_frames.unsqueeze(-1)], dim=-1)
        shared = self.shared(x)
        score = self.score_head(shared).squeeze(-1)
        pherr_logit = self.pherr_head(shared).squeeze(-1)
        return score, pherr_logit


# ============================================================
# Main Pipeline
# ============================================================
HF_REPO_ID = "Jianshu001/wavlm-phoneme-scorer"
HF_CHECKPOINT_FILENAME = "wavlm_finetuned.pt"


class PronunciationAssessorV2:
    """Phoneme-level pronunciation assessment using fine-tuned WavLM backbone."""

    def __init__(self, checkpoint_path=None, device=None, pherr_threshold=DEFAULT_PHERR_THRESHOLD):
        self.device = device or torch.device("cuda:0" if torch.cuda.is_available() else "cpu")
        self.pherr_threshold = pherr_threshold
        self._checkpoint_path = checkpoint_path or self._resolve_checkpoint()
        self._backbone = None
        self._scorer = None
        self._fe_backbone = None
        self._ctc_model = None
        self._fe_ctc = None
        self._vocab = None
        self._blank_idx = None
        self._g2p = None

    @staticmethod
    def _resolve_checkpoint():
        """Try local file first, then download from HuggingFace."""
        local_path = Path(__file__).parent / HF_CHECKPOINT_FILENAME
        if local_path.exists():
            return str(local_path)
        print(f"Downloading model from huggingface.co/{HF_REPO_ID}...", file=sys.stderr)
        return hf_hub_download(repo_id=HF_REPO_ID, filename=HF_CHECKPOINT_FILENAME)

    @classmethod
    def from_pretrained(cls, repo_id=None, device=None, pherr_threshold=DEFAULT_PHERR_THRESHOLD):
        """
        Load model from HuggingFace Hub.

        Args:
            repo_id: HuggingFace repo (default: Jianshu001/wavlm-phoneme-scorer)
            device: torch device
            pherr_threshold: error detection threshold

        Example:
            assessor = PronunciationAssessorV2.from_pretrained()
            result = assessor.assess("audio.mp3", "Hello, Peter.")
        """
        repo = repo_id or HF_REPO_ID
        print(f"Downloading model from huggingface.co/{repo}...", file=sys.stderr)
        checkpoint_path = hf_hub_download(repo_id=repo, filename=HF_CHECKPOINT_FILENAME)
        return cls(checkpoint_path=checkpoint_path, device=device, pherr_threshold=pherr_threshold)

    def _load_models(self):
        if self._backbone is not None:
            return

        print("Loading WavLM backbone + scoring head...", file=sys.stderr)

        # Load checkpoint
        ckpt = torch.load(self._checkpoint_path, map_location="cpu", weights_only=False)
        state_dict = ckpt["model_state"]

        # Separate backbone and head state dicts
        backbone_state = {}
        head_state = {}
        for k, v in state_dict.items():
            if k.startswith("backbone."):
                backbone_state[k[len("backbone."):]] = v
            else:
                head_state[k] = v

        # Load WavLM backbone
        self._backbone = WavLMModel.from_pretrained(
            "microsoft/wavlm-large",
            output_hidden_states=False,
            mask_time_prob=0.0,
        )
        self._backbone.load_state_dict(backbone_state, strict=False)
        self._backbone.to(self.device)
        self._backbone.eval()

        self._fe_backbone = Wav2Vec2FeatureExtractor.from_pretrained("microsoft/wavlm-large")

        # Load scoring head
        self._scorer = PhoneScorerHead()
        self._scorer.load_state_dict(head_state)
        self._scorer.to(self.device)
        self._scorer.eval()

        # Load CTC model for alignment
        print("Loading CTC alignment model...", file=sys.stderr)
        ctc_name = "facebook/wav2vec2-xlsr-53-espeak-cv-ft"
        self._ctc_model = Wav2Vec2ForCTC.from_pretrained(ctc_name).to(self.device)
        self._ctc_model.eval()
        self._fe_ctc = Wav2Vec2FeatureExtractor.from_pretrained(ctc_name)

        vocab_path = hf_hub_download(ctc_name, "vocab.json")
        with open(vocab_path) as f:
            self._vocab = json.load(f)
        self._blank_idx = self._vocab.get("<pad>", 0)

        self._g2p = G2p()
        print(f"Models loaded (device={self.device})", file=sys.stderr)

    # --------------------------------------------------------
    # Audio
    # --------------------------------------------------------
    @staticmethod
    def load_audio(audio_path):
        waveform, sr = torchaudio.load(audio_path)
        if waveform.shape[0] > 1:
            waveform = waveform.mean(0, keepdim=True)
        if sr != SAMPLE_RATE:
            waveform = F_audio.resample(waveform, sr, SAMPLE_RATE)
        return waveform

    # --------------------------------------------------------
    # G2P
    # --------------------------------------------------------
    def _text_to_phonemes(self, text):
        words = re.sub(r"[^\w' ]", " ", text).split()
        result = []
        for word_idx, word in enumerate(words):
            phones_raw = self._g2p(word)
            for ph in phones_raw:
                if ph == " ":
                    continue
                clean = re.sub(r"\d", "", ph).lower()
                if clean:
                    result.append({"phone": clean, "word": word, "word_idx": word_idx})
        return result

    def _arpabet_to_model_idx(self, ph):
        for ipa in ARPABET_TO_IPA.get(ph, []):
            if ipa in self._vocab:
                return self._vocab[ipa]
        return self._vocab.get(ph, -1)

    # --------------------------------------------------------
    # Viterbi forced alignment
    # --------------------------------------------------------
    @staticmethod
    def _viterbi_align(emissions, phone_indices, blank_idx):
        T, C = emissions.shape
        S = len(phone_indices)
        if S == 0 or T < S:
            return []

        extended = [blank_idx]
        for p in phone_indices:
            extended.append(p)
            extended.append(blank_idx)
        S_ext = len(extended)

        NEG_INF = float("-inf")
        dp = np.full((T, S_ext), NEG_INF, dtype=np.float64)
        bp = np.zeros((T, S_ext), dtype=np.int32)
        dp[0][0] = emissions[0, extended[0]].item()
        if S_ext > 1:
            dp[0][1] = emissions[0, extended[1]].item()

        for t in range(1, T):
            for s in range(S_ext):
                emit = emissions[t, extended[s]].item()
                best, best_s = dp[t - 1][s], s
                if s > 0 and dp[t - 1][s - 1] > best:
                    best, best_s = dp[t - 1][s - 1], s - 1
                if s > 1 and extended[s] != blank_idx and extended[s] != extended[s - 2]:
                    if dp[t - 1][s - 2] > best:
                        best, best_s = dp[t - 1][s - 2], s - 2
                dp[t][s] = best + emit
                bp[t][s] = best_s

        s = S_ext - 1 if (S_ext >= 2 and dp[T - 1][S_ext - 1] >= dp[T - 1][S_ext - 2]) else max(S_ext - 2, 0)
        path = []
        for t in range(T - 1, -1, -1):
            path.append((t, extended[s]))
            s = bp[t][s]
        path.reverse()
        return path

    # --------------------------------------------------------
    # Core: align + extract features + score
    # --------------------------------------------------------
    @torch.no_grad()
    def _score_phonemes(self, waveform, phone_indices):
        """
        Given waveform and expected phone indices, returns per-phoneme scores.
        Steps:
            1. CTC emissions → Viterbi alignment → frame segments
            2. WavLM backbone → hidden states per segment
            3. GOP from CTC emissions
            4. MLP head → score + pherr probability
        """
        wav_np = waveform.squeeze(0).numpy()

        # CTC emissions for alignment
        ctc_inputs = self._fe_ctc(wav_np, sampling_rate=SAMPLE_RATE, return_tensors="pt", padding=True)
        ctc_logits = self._ctc_model(ctc_inputs.input_values.to(self.device)).logits
        emissions = torch.log_softmax(ctc_logits, dim=-1).squeeze(0).cpu()

        # Viterbi alignment
        path = self._viterbi_align(emissions, phone_indices, self._blank_idx)
        if not path:
            return None

        # Group frames by phoneme
        segments = []
        cur_tok, cur_frames = None, []
        for f, tok in path:
            if tok == self._blank_idx:
                if cur_tok is not None:
                    segments.append((cur_tok, cur_frames))
                    cur_tok, cur_frames = None, []
                continue
            if tok != cur_tok:
                if cur_tok is not None:
                    segments.append((cur_tok, cur_frames))
                cur_tok, cur_frames = tok, [f]
            else:
                cur_frames.append(f)
        if cur_tok is not None:
            segments.append((cur_tok, cur_frames))

        # WavLM backbone hidden states
        bb_inputs = self._fe_backbone(wav_np, sampling_rate=SAMPLE_RATE, return_tensors="pt", padding=True)
        hidden = self._backbone(bb_inputs.input_values.to(self.device)).last_hidden_state.squeeze(0)  # (T_h, 1024)
        T_h = hidden.shape[0]
        T_ctc = emissions.shape[0]
        scale = T_h / T_ctc

        # Per-phoneme: pool hidden states, compute GOP, run MLP
        results = []
        for i, expected_idx in enumerate(phone_indices):
            if i >= len(segments):
                results.append({"gop": -20.0, "score": 0.0, "pherr_prob": 1.0})
                continue

            _, frames = segments[i]

            # Pool hidden states
            h_start = max(0, int(min(frames) * scale))
            h_end = min(T_h, int((max(frames) + 1) * scale))
            if h_end <= h_start:
                h_end = h_start + 1
            h_pooled = hidden[h_start:h_end].mean(dim=0)  # (1024,)

            # GOP
            seg_em = emissions[frames]
            target_lp = seg_em[:, expected_idx].mean().item()
            mask = torch.ones(emissions.shape[1], dtype=torch.bool)
            mask[self._blank_idx] = False
            mask[expected_idx] = False
            best_other = seg_em[:, mask].max(dim=-1).values.mean().item()
            gop = target_lp - best_other

            results.append({
                "h": h_pooled,
                "gop": gop,
                "n_frames": len(frames),
            })

        # Batch MLP inference
        valid_indices = [i for i, r in enumerate(results) if "h" in r]
        if valid_indices:
            h_batch = torch.stack([results[i]["h"] for i in valid_indices]).to(self.device)
            gop_batch = torch.tensor([results[i]["gop"] for i in valid_indices],
                                     dtype=torch.float32).to(self.device)
            nf_batch = torch.tensor([results[i]["n_frames"] for i in valid_indices],
                                    dtype=torch.float32).to(self.device)

            # Need phone_ids for the valid phonemes
            pid_batch = torch.tensor([PHONE_TO_ID.get(
                # We need to pass phone names - will be set from caller
                "_placeholder_", 0) for _ in valid_indices],
                dtype=torch.long).to(self.device)

            # Return raw data for batch processing in assess()
            return results, valid_indices, h_batch, gop_batch, nf_batch

        return results, [], None, None, None

    # --------------------------------------------------------
    # Public API
    # --------------------------------------------------------
    def assess(self, audio_path, text):
        """
        Assess pronunciation of an audio file against reference text.

        Returns dict with overall_score, words (with per-phoneme scores and errors).
        """
        self._load_models()

        # G2P
        phone_info = self._text_to_phonemes(text)
        if not phone_info:
            return {"text": text, "overall_score": 0, "words": [], "error": "No phonemes extracted"}

        # Map to model indices
        indices = []
        valid_info = []
        for pi in phone_info:
            idx = self._arpabet_to_model_idx(pi["phone"])
            if idx >= 0:
                indices.append(idx)
                valid_info.append(pi)
        if not indices:
            return {"text": text, "overall_score": 0, "words": [], "error": "No phonemes mapped"}

        # Load audio & score
        waveform = self.load_audio(audio_path)
        result = self._score_phonemes(waveform, indices)
        if result is None:
            return {"text": text, "overall_score": 0, "words": [], "error": "Alignment failed"}

        raw_results, valid_idx, h_batch, gop_batch, nf_batch = result

        # Run MLP scoring
        if h_batch is not None:
            pid_list = [PHONE_TO_ID.get(valid_info[i]["phone"], 0) for i in valid_idx]
            pid_batch = torch.tensor(pid_list, dtype=torch.long).to(self.device)

            pred_score, pred_pherr_logit = self._scorer(h_batch, pid_batch, gop_batch, nf_batch)
            pred_score = pred_score.detach().cpu().numpy()
            pred_pherr = torch.sigmoid(pred_pherr_logit).detach().cpu().numpy()

            for j, i in enumerate(valid_idx):
                raw_results[i]["score"] = float(np.clip(pred_score[j], 0, 100))
                raw_results[i]["pherr_prob"] = float(pred_pherr[j])

        # Assemble per-phoneme results
        phoneme_results = []
        for i, (info, raw) in enumerate(zip(valid_info, raw_results)):
            score = raw.get("score", 0.0)
            pherr_prob = raw.get("pherr_prob", 1.0)
            phoneme_results.append({
                "phone": info["phone"],
                "word": info["word"],
                "word_idx": info["word_idx"],
                "score": round(score, 1),
                "gop": round(raw["gop"], 3),
                "pherr_prob": round(pherr_prob, 3),
                "error": pherr_prob >= self.pherr_threshold,
            })

        # Group by word
        words_dict = {}
        for pr in phoneme_results:
            widx = pr["word_idx"]
            if widx not in words_dict:
                words_dict[widx] = {"word": pr["word"], "phonemes": []}
            words_dict[widx]["phonemes"].append({
                "phone": pr["phone"],
                "score": pr["score"],
                "gop": pr["gop"],
                "pherr_prob": pr["pherr_prob"],
                "error": pr["error"],
            })

        # Word-level scores
        words_list = []
        for widx in sorted(words_dict.keys()):
            wd = words_dict[widx]
            scores = [p["score"] for p in wd["phonemes"]]
            n_errors = sum(1 for p in wd["phonemes"] if p["error"])
            mean_score = np.mean(scores)
            words_list.append({
                "word": wd["word"],
                "score": round(float(mean_score), 1),
                "n_errors": n_errors,
                "n_phonemes": len(wd["phonemes"]),
                "has_error": n_errors > 0,
                "phonemes": wd["phonemes"],
            })

        # Overall score
        all_scores = [pr["score"] for pr in phoneme_results]
        overall_score = np.mean(all_scores)
        n_total_errors = sum(1 for pr in phoneme_results if pr["error"])

        return {
            "text": text,
            "overall_score": round(float(overall_score), 1),
            "n_phonemes": len(phoneme_results),
            "n_errors": n_total_errors,
            "error_rate": round(n_total_errors / len(phoneme_results) * 100, 1),
            "words": words_list,
        }

    def assess_batch(self, items, show_progress=True):
        """Assess a list of (audio_path, text) pairs."""
        self._load_models()
        results = []
        for i, (audio_path, text) in enumerate(items):
            try:
                result = self.assess(audio_path, text)
                results.append(result)
            except Exception as e:
                results.append({"text": text, "overall_score": 0, "error": str(e)})
            if show_progress and (i + 1) % 50 == 0:
                print(f"  Processed {i + 1}/{len(items)}...", file=sys.stderr)
        return results


# ============================================================
# Pretty print
# ============================================================
def print_result(result):
    print(f"\n{'='*60}")
    print(f"Text: \"{result['text']}\"")
    print(f"Overall Score: {result['overall_score']:.1f}/100  "
          f"(errors: {result.get('n_errors', '?')}/{result.get('n_phonemes', '?')})")
    print(f"{'='*60}")

    for wd in result.get("words", []):
        status = "\u2717" if wd["has_error"] else "\u2713"
        print(f"\n  {status} {wd['word']:<15s}  score={wd['score']:5.1f}  "
              f"errors={wd['n_errors']}/{wd['n_phonemes']}")

        for ph in wd["phonemes"]:
            marker = " \u2190 ERROR" if ph["error"] else ""
            print(f"      /{ph['phone']:<4s}/  score={ph['score']:5.1f}  "
                  f"GOP={ph['gop']:+6.2f}  pherr={ph['pherr_prob']:.2f}{marker}")


# ============================================================
# CLI
# ============================================================
def main():
    parser = argparse.ArgumentParser(description="Voice Correction Pipeline v2.0 (WavLM Fine-tuned)")
    parser.add_argument("--audio", type=str, help="Path to audio file")
    parser.add_argument("--text", type=str, help="Reference text")
    parser.add_argument("--checkpoint", type=str, default=None,
                        help="Path to fine-tuned model checkpoint (default: wavlm_finetuned.pt)")
    parser.add_argument("--threshold", type=float, default=DEFAULT_PHERR_THRESHOLD,
                        help=f"Pherr probability threshold (default: {DEFAULT_PHERR_THRESHOLD})")
    parser.add_argument("--batch", action="store_true", help="Batch mode")
    parser.add_argument("--input", type=str, help="Input xlsx file (batch mode)")
    parser.add_argument("--audio-dir", type=str, help="Audio directory (batch mode)")
    parser.add_argument("--output", type=str, help="Output JSON file")
    parser.add_argument("--limit", type=int, default=0, help="Limit number of samples (0=all)")
    parser.add_argument("--evaluate", action="store_true", help="Evaluate against ground truth")
    parser.add_argument("--json", action="store_true", help="Output raw JSON")
    parser.add_argument("--device", type=str, default=None, help="Device (cuda:0, cpu)")
    args = parser.parse_args()

    device = torch.device(args.device) if args.device else None
    assessor = PronunciationAssessorV2(
        checkpoint_path=args.checkpoint, device=device, pherr_threshold=args.threshold
    )

    if args.batch:
        if not args.input:
            parser.error("--input required for batch mode")

        import openpyxl
        audio_dir = Path(args.audio_dir) if args.audio_dir else Path(args.input).parent / "audio_files"

        wb = openpyxl.load_workbook(args.input, read_only=True)
        ws = wb.active
        items = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            fname, content = row[0], row[1]
            if not fname or not content:
                continue
            audio_path = audio_dir / fname
            if audio_path.exists():
                items.append((str(audio_path), content))
        wb.close()

        if args.limit > 0:
            items = items[:args.limit]

        print(f"Processing {len(items)} files...", file=sys.stderr)
        results = assessor.assess_batch(items)

        if args.output:
            with open(args.output, "w") as f:
                json.dump(results, f, indent=2, ensure_ascii=False)
            print(f"\nResults saved to {args.output}", file=sys.stderr)
        else:
            for r in results:
                if args.json:
                    print(json.dumps(r, ensure_ascii=False))
                else:
                    print_result(r)

        if args.evaluate:
            _run_evaluation(results, args.input)

    else:
        if not args.audio or not args.text:
            parser.error("--audio and --text are required")

        result = assessor.assess(args.audio, args.text)
        if args.json:
            print(json.dumps(result, indent=2, ensure_ascii=False))
        else:
            print_result(result)


def _run_evaluation(results, xlsx_path):
    """Compare pipeline output against ground truth."""
    import openpyxl
    from sklearn.metrics import precision_recall_fscore_support, roc_auc_score
    from scipy import stats

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active

    gt_records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        raw = row[2]
        if not raw:
            continue
        s = raw
        for _ in range(5):
            s = s.replace("\\\\", "\\")
        s = s.replace('\\"', '"')
        acc = re.search(r'"accuracy":\s*([\d.]+)', s)
        phones = re.findall(
            r'"char":"([^"]+)","ph2alpha":"([^"]*)".*?"pherr":(\d+).*?"score":([\d.]+)', s
        )
        gt_records.append({
            "accuracy": float(acc.group(1)) if acc else 0,
            "phones": [{"pherr": int(e), "score": float(sc)} for _, _, e, sc in phones],
        })
    wb.close()

    # Overall accuracy correlation
    n = min(len(gt_records), len(results))
    gt_acc = np.array([r["accuracy"] for r in gt_records[:n]])
    pred_acc = np.array([r.get("overall_score", 0) for r in results[:n]])
    mae_acc = np.mean(np.abs(gt_acc - pred_acc))
    corr_acc, _ = stats.pearsonr(gt_acc, pred_acc)

    # Phoneme-level metrics
    all_gt_pherr, all_pred_pherr = [], []
    all_gt_score, all_pred_score = [], []

    for i in range(n):
        if "error" in results[i]:
            continue
        gt_phones = gt_records[i]["phones"]
        pred_words = results[i].get("words", [])
        pred_phones = []
        for w in pred_words:
            pred_phones.extend(w.get("phonemes", []))

        m = min(len(gt_phones), len(pred_phones))
        for j in range(m):
            all_gt_pherr.append(gt_phones[j]["pherr"])
            all_pred_pherr.append(pred_phones[j]["pherr_prob"])
            all_gt_score.append(gt_phones[j]["score"])
            all_pred_score.append(pred_phones[j]["score"])

    gt_pherr = np.array(all_gt_pherr)
    pred_pherr = np.array(all_pred_pherr)
    gt_score = np.array(all_gt_score)
    pred_score = np.array(all_pred_score)

    auc = roc_auc_score(gt_pherr, pred_pherr) if len(np.unique(gt_pherr)) > 1 else 0

    best_f1, best_th = 0, 0.5
    for th in np.arange(0.1, 0.9, 0.05):
        pb = (pred_pherr >= th).astype(int)
        _, _, f1, _ = precision_recall_fscore_support(gt_pherr, pb, average="binary", zero_division=0)
        if f1 > best_f1:
            best_f1, best_th = f1, th

    pb = (pred_pherr >= best_th).astype(int)
    prec, rec, f1, _ = precision_recall_fscore_support(gt_pherr, pb, average="binary")

    corr_phone, _ = stats.pearsonr(gt_score, pred_score)
    mae_phone = np.mean(np.abs(gt_score - pred_score))

    print(f"\n{'='*60}", file=sys.stderr)
    print(f"EVALUATION ({n} samples, {len(gt_pherr)} phonemes)", file=sys.stderr)
    print(f"{'='*60}", file=sys.stderr)
    print(f"  Overall accuracy MAE:     {mae_acc:.2f}", file=sys.stderr)
    print(f"  Overall accuracy Pearson: {corr_acc:.3f}", file=sys.stderr)
    print(f"\n  Phoneme error AUC-ROC:    {auc:.3f}", file=sys.stderr)
    print(f"  Phoneme error F1:         {f1:.3f} (threshold={best_th:.2f})", file=sys.stderr)
    print(f"  Phoneme error Precision:  {prec:.3f}", file=sys.stderr)
    print(f"  Phoneme error Recall:     {rec:.3f}", file=sys.stderr)
    print(f"\n  Phone score Pearson:      {corr_phone:.3f}", file=sys.stderr)
    print(f"  Phone score MAE:          {mae_phone:.2f}", file=sys.stderr)


if __name__ == "__main__":
    main()
