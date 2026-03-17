"""
Voice Correction Pipeline v1.0
==============================
Given a reference text and an audio recording, outputs phoneme-level
pronunciation assessment: which phonemes are correct and which are wrong.

Usage:
    # Single file
    python pipeline.py --audio audio_files/example.mp3 --text "Hello, Peter."

    # Batch mode (process eval_log.xlsx)
    python pipeline.py --batch --input eval_log.xlsx --audio-dir audio_files/ --output results.json --limit 20

    # Evaluate against ground truth
    python pipeline.py --batch --input eval_log.xlsx --audio-dir audio_files/ --evaluate
"""

import argparse
import json
import re
import sys
import warnings
from pathlib import Path

import numpy as np
import torch
import torchaudio
import torchaudio.functional as F_audio
from g2p_en import G2p
from huggingface_hub import hf_hub_download
from transformers import Wav2Vec2FeatureExtractor, Wav2Vec2ForCTC

warnings.filterwarnings("ignore")

# ============================================================
# ARPAbet ↔ IPA mapping
# ============================================================
ARPABET_TO_IPA = {
    "aa": ["ɑː", "ɑ", "ɒ", "a"], "ae": ["æ"], "ah": ["ʌ", "ə", "ɐ"],
    "ao": ["ɔː", "ɔ", "ɒ"], "aw": ["aʊ"], "ax": ["ə", "ɐ", "ʌ"], "ay": ["aɪ"],
    "b": ["b"], "ch": ["tʃ"], "d": ["d"], "dh": ["ð"],
    "eh": ["ɛ", "e"], "er": ["ɜː", "ɝ", "ɚ", "ɜ"], "ey": ["eɪ"],
    "f": ["f"], "g": ["ɡ", "g"], "hh": ["h"],
    "ih": ["ɪ", "ᵻ"], "iy": ["iː", "i"],
    "ir": ["ɪɹ"], "jh": ["dʒ"], "k": ["k"], "l": ["l"],
    "m": ["m"], "n": ["n"], "ng": ["ŋ"],
    "ow": ["oʊ", "o", "əʊ"], "oy": ["ɔɪ"],
    "p": ["p"], "r": ["ɹ", "r"], "s": ["s"], "sh": ["ʃ"],
    "t": ["t"], "th": ["θ"], "uh": ["ʊ"], "uw": ["uː", "u"],
    "ur": ["ʊɹ"], "v": ["v"], "w": ["w"], "y": ["j"],
    "z": ["z"], "zh": ["ʒ"],
    "ar": ["ɑːɹ"], "oo": ["ʊ", "uː"], "dr": ["dɹ"], "tr": ["tɹ"],
    "ts": ["ts"], "dz": ["dz"],
}

# Default GOP threshold for error detection
DEFAULT_THRESHOLD = -2.5
SAMPLE_RATE = 16000


# ============================================================
# Model Manager (lazy loading)
# ============================================================
class PronunciationAssessor:
    """Phoneme-level pronunciation assessment using GOP scoring."""

    def __init__(self, device=None, threshold=DEFAULT_THRESHOLD):
        self.device = device or torch.device("cuda:0" if torch.cuda.is_available() else "cpu")
        self.threshold = threshold
        self._model = None
        self._feat_extractor = None
        self._vocab = None
        self._idx2phone = None
        self._g2p = None
        self._blank_idx = None

    def _load_model(self):
        if self._model is not None:
            return
        print("Loading wav2vec2 phoneme model...", file=sys.stderr)
        model_name = "facebook/wav2vec2-xlsr-53-espeak-cv-ft"
        self._feat_extractor = Wav2Vec2FeatureExtractor.from_pretrained(model_name)
        self._model = Wav2Vec2ForCTC.from_pretrained(model_name).to(self.device)
        self._model.eval()

        vocab_path = hf_hub_download(model_name, "vocab.json")
        with open(vocab_path) as f:
            self._vocab = json.load(f)
        self._idx2phone = {v: k for k, v in self._vocab.items()}
        self._blank_idx = self._vocab.get("<pad>", 0)
        self._g2p = G2p()
        print(f"Model loaded ({len(self._vocab)} phonemes, device={self.device})", file=sys.stderr)

    # --------------------------------------------------------
    # Audio loading
    # --------------------------------------------------------
    @staticmethod
    def load_audio(audio_path: str) -> torch.Tensor:
        waveform, sr = torchaudio.load(audio_path)
        if waveform.shape[0] > 1:
            waveform = waveform.mean(0, keepdim=True)
        if sr != SAMPLE_RATE:
            waveform = F_audio.resample(waveform, sr, SAMPLE_RATE)
        return waveform

    # --------------------------------------------------------
    # Phoneme emissions
    # --------------------------------------------------------
    @torch.no_grad()
    def _get_emissions(self, waveform: torch.Tensor) -> torch.Tensor:
        inputs = self._feat_extractor(
            waveform.squeeze(0).numpy(), sampling_rate=SAMPLE_RATE,
            return_tensors="pt", padding=True,
        )
        logits = self._model(inputs.input_values.to(self.device)).logits
        return torch.log_softmax(logits, dim=-1).squeeze(0).cpu()

    # --------------------------------------------------------
    # G2P: text → ARPAbet phoneme list with word boundaries
    # --------------------------------------------------------
    def _text_to_phonemes(self, text: str) -> list:
        """Returns list of dicts: {phone, word, word_idx}."""
        words = re.sub(r"[^\w' ]", " ", text).split()
        result = []
        for word_idx, word in enumerate(words):
            phones_raw = self._g2p(word)
            for ph in phones_raw:
                if ph == " ":
                    continue
                clean = re.sub(r"\d", "", ph).lower()
                if clean:
                    result.append({"phone": clean, "word": word, "word_idx": word_idx})
        return result

    def _arpabet_to_model_idx(self, ph: str) -> int:
        candidates = ARPABET_TO_IPA.get(ph, [])
        for ipa in candidates:
            if ipa in self._vocab:
                return self._vocab[ipa]
        if ph in self._vocab:
            return self._vocab[ph]
        return -1

    # --------------------------------------------------------
    # Viterbi forced alignment
    # --------------------------------------------------------
    @staticmethod
    def _viterbi_align(emissions: torch.Tensor, phone_indices: list, blank_idx: int) -> list:
        T, C = emissions.shape
        S = len(phone_indices)
        if S == 0 or T < S:
            return []

        extended = [blank_idx]
        for p in phone_indices:
            extended.append(p)
            extended.append(blank_idx)
        S_ext = len(extended)

        NEG_INF = float("-inf")
        dp = np.full((T, S_ext), NEG_INF, dtype=np.float64)
        bp = np.zeros((T, S_ext), dtype=np.int32)
        dp[0][0] = emissions[0, extended[0]].item()
        if S_ext > 1:
            dp[0][1] = emissions[0, extended[1]].item()

        for t in range(1, T):
            for s in range(S_ext):
                emit = emissions[t, extended[s]].item()
                best, best_s = dp[t - 1][s], s
                if s > 0 and dp[t - 1][s - 1] > best:
                    best, best_s = dp[t - 1][s - 1], s - 1
                if s > 1 and extended[s] != blank_idx and extended[s] != extended[s - 2]:
                    if dp[t - 1][s - 2] > best:
                        best, best_s = dp[t - 1][s - 2], s - 2
                dp[t][s] = best + emit
                bp[t][s] = best_s

        s = S_ext - 1 if (S_ext >= 2 and dp[T - 1][S_ext - 1] >= dp[T - 1][S_ext - 2]) else max(S_ext - 2, 0)
        path = []
        for t in range(T - 1, -1, -1):
            path.append((t, extended[s]))
            s = bp[t][s]
        path.reverse()
        return path

    # --------------------------------------------------------
    # GOP computation
    # --------------------------------------------------------
    def _compute_gop(self, emissions: torch.Tensor, phone_indices: list) -> list:
        path = self._viterbi_align(emissions, phone_indices, self._blank_idx)
        if not path:
            return [{"gop": -20.0, "log_post": -20.0}] * len(phone_indices)

        # Group frames by phoneme
        segments = []
        cur_tok, cur_frames = None, []
        for f, tok in path:
            if tok == self._blank_idx:
                if cur_tok is not None:
                    segments.append((cur_tok, cur_frames))
                    cur_tok, cur_frames = None, []
                continue
            if tok != cur_tok:
                if cur_tok is not None:
                    segments.append((cur_tok, cur_frames))
                cur_tok, cur_frames = tok, [f]
            else:
                cur_frames.append(f)
        if cur_tok is not None:
            segments.append((cur_tok, cur_frames))

        C = emissions.shape[1]
        results = []
        for i, expected_idx in enumerate(phone_indices):
            if i >= len(segments):
                results.append({"gop": -20.0, "log_post": -20.0})
                continue
            _, frames = segments[i]
            frame_em = emissions[frames]
            target_lp = frame_em[:, expected_idx].mean().item()

            mask = torch.ones(C, dtype=torch.bool)
            mask[self._blank_idx] = False
            mask[expected_idx] = False
            best_other = frame_em[:, mask].max(dim=-1).values.mean().item() if mask.any() else -30.0

            results.append({
                "gop": round(target_lp - best_other, 3),
                "log_post": round(target_lp, 3),
            })
        return results

    # --------------------------------------------------------
    # Public API
    # --------------------------------------------------------
    def assess(self, audio_path: str, text: str) -> dict:
        """
        Assess pronunciation of an audio file against reference text.

        Returns:
            {
                "text": "Hello, Peter.",
                "overall_score": 75.0,
                "words": [
                    {
                        "word": "Hello",
                        "score": 80.0,
                        "phonemes": [
                            {"phone": "hh", "gop": 4.9, "error": false},
                            {"phone": "ah", "gop": 0.6, "error": false},
                            ...
                        ]
                    },
                    ...
                ]
            }
        """
        self._load_model()

        # G2P
        phone_info = self._text_to_phonemes(text)
        if not phone_info:
            return {"text": text, "overall_score": 0, "words": [], "error": "No phonemes extracted"}

        # Map to model indices
        indices = []
        valid_info = []
        for pi in phone_info:
            idx = self._arpabet_to_model_idx(pi["phone"])
            if idx >= 0:
                indices.append(idx)
                valid_info.append(pi)
        if not indices:
            return {"text": text, "overall_score": 0, "words": [], "error": "No phonemes mapped to model vocab"}

        # Get emissions & compute GOP
        waveform = self.load_audio(audio_path)
        emissions = self._get_emissions(waveform)
        gop_results = self._compute_gop(emissions, indices)

        # Assemble per-phoneme results
        phoneme_results = []
        for info, gop_r in zip(valid_info, gop_results):
            phoneme_results.append({
                "phone": info["phone"],
                "word": info["word"],
                "word_idx": info["word_idx"],
                "gop": gop_r["gop"],
                "log_posterior": gop_r["log_post"],
                "error": gop_r["gop"] < self.threshold,
            })

        # Group by word
        words_dict = {}
        for pr in phoneme_results:
            widx = pr["word_idx"]
            if widx not in words_dict:
                words_dict[widx] = {"word": pr["word"], "phonemes": []}
            words_dict[widx]["phonemes"].append({
                "phone": pr["phone"],
                "gop": pr["gop"],
                "error": pr["error"],
            })

        # Word-level scores
        words_list = []
        for widx in sorted(words_dict.keys()):
            wd = words_dict[widx]
            gops = [p["gop"] for p in wd["phonemes"]]
            n_errors = sum(1 for p in wd["phonemes"] if p["error"])
            # Score: map mean GOP to 0-100 range
            mean_gop = np.mean(gops)
            score = max(0, min(100, (mean_gop + 10) * 10))
            words_list.append({
                "word": wd["word"],
                "score": round(score, 1),
                "mean_gop": round(float(mean_gop), 2),
                "n_errors": n_errors,
                "n_phonemes": len(wd["phonemes"]),
                "has_error": n_errors > 0,
                "phonemes": wd["phonemes"],
            })

        # Overall score
        all_gops = [pr["gop"] for pr in phoneme_results]
        overall_gop = np.mean(all_gops)
        overall_score = max(0, min(100, (overall_gop + 10) * 10))
        n_total_errors = sum(1 for pr in phoneme_results if pr["error"])

        return {
            "text": text,
            "overall_score": round(float(overall_score), 1),
            "overall_gop": round(float(overall_gop), 2),
            "n_phonemes": len(phoneme_results),
            "n_errors": n_total_errors,
            "error_rate": round(n_total_errors / len(phoneme_results) * 100, 1),
            "words": words_list,
        }

    def assess_batch(self, items: list, show_progress=True) -> list:
        """Assess a list of (audio_path, text) pairs."""
        self._load_model()
        results = []
        for i, (audio_path, text) in enumerate(items):
            try:
                result = self.assess(audio_path, text)
                results.append(result)
            except Exception as e:
                results.append({"text": text, "overall_score": 0, "error": str(e)})
            if show_progress and (i + 1) % 50 == 0:
                print(f"  Processed {i + 1}/{len(items)}...", file=sys.stderr)
        return results


# ============================================================
# Pretty print
# ============================================================
def print_result(result: dict):
    """Print assessment result in a readable format."""
    print(f"\n{'='*60}")
    print(f"Text: \"{result['text']}\"")
    print(f"Overall Score: {result['overall_score']:.1f}/100  "
          f"(errors: {result.get('n_errors', '?')}/{result.get('n_phonemes', '?')})")
    print(f"{'='*60}")

    for wd in result.get("words", []):
        status = "✗" if wd["has_error"] else "✓"
        print(f"\n  {status} {wd['word']:<15s}  score={wd['score']:5.1f}  "
              f"errors={wd['n_errors']}/{wd['n_phonemes']}")

        for ph in wd["phonemes"]:
            marker = " ← ERROR" if ph["error"] else ""
            print(f"      /{ph['phone']:<4s}/  GOP={ph['gop']:+6.2f}{marker}")


# ============================================================
# CLI
# ============================================================
def main():
    parser = argparse.ArgumentParser(description="Voice Correction Pipeline v1.0")
    parser.add_argument("--audio", type=str, help="Path to audio file")
    parser.add_argument("--text", type=str, help="Reference text")
    parser.add_argument("--threshold", type=float, default=DEFAULT_THRESHOLD,
                        help=f"GOP threshold for error detection (default: {DEFAULT_THRESHOLD})")
    parser.add_argument("--batch", action="store_true", help="Batch mode")
    parser.add_argument("--input", type=str, help="Input xlsx file (batch mode)")
    parser.add_argument("--audio-dir", type=str, help="Audio directory (batch mode)")
    parser.add_argument("--output", type=str, help="Output JSON file")
    parser.add_argument("--limit", type=int, default=0, help="Limit number of samples (0=all)")
    parser.add_argument("--evaluate", action="store_true", help="Evaluate against ground truth")
    parser.add_argument("--json", action="store_true", help="Output raw JSON instead of pretty print")
    parser.add_argument("--device", type=str, default=None, help="Device (cuda:0, cpu)")
    args = parser.parse_args()

    device = torch.device(args.device) if args.device else None
    assessor = PronunciationAssessor(device=device, threshold=args.threshold)

    if args.batch:
        # Batch mode
        if not args.input:
            parser.error("--input required for batch mode")

        import openpyxl
        audio_dir = Path(args.audio_dir) if args.audio_dir else Path(args.input).parent / "audio_files"

        wb = openpyxl.load_workbook(args.input)
        ws = wb.active
        items = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            fname, content = row[0], row[1]
            if not fname or not content:
                continue
            audio_path = audio_dir / fname
            if audio_path.exists():
                items.append((str(audio_path), content))

        if args.limit > 0:
            items = items[:args.limit]

        print(f"Processing {len(items)} files...", file=sys.stderr)
        results = assessor.assess_batch(items)

        if args.output:
            with open(args.output, "w") as f:
                json.dump(results, f, indent=2, ensure_ascii=False)
            print(f"\nResults saved to {args.output}", file=sys.stderr)
        else:
            for r in results:
                if args.json:
                    print(json.dumps(r, ensure_ascii=False))
                else:
                    print_result(r)

        if args.evaluate:
            _run_evaluation(results, args.input)

    else:
        # Single file mode
        if not args.audio or not args.text:
            parser.error("--audio and --text are required")

        result = assessor.assess(args.audio, args.text)
        if args.json:
            print(json.dumps(result, indent=2, ensure_ascii=False))
        else:
            print_result(result)


def _run_evaluation(results, xlsx_path):
    """Compare pipeline output against ground truth from eval_log."""
    import openpyxl
    from scipy import stats

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    gt_accs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        raw = row[2]
        if not raw:
            continue
        s = raw
        for _ in range(5):
            s = s.replace("\\\\", "\\")
        s = s.replace('\\"', '"')
        acc = re.search(r'"accuracy":\s*([\d.]+)', s)
        gt_accs.append(float(acc.group(1)) if acc else 0)

    pred_scores = [r.get("overall_score", 0) for r in results]
    n = min(len(gt_accs), len(pred_scores))
    gt_arr = np.array(gt_accs[:n])
    pred_arr = np.array(pred_scores[:n])

    mae = np.mean(np.abs(gt_arr - pred_arr))
    corr, _ = stats.pearsonr(gt_arr, pred_arr)
    sp, _ = stats.spearmanr(gt_arr, pred_arr)

    print(f"\n{'='*60}", file=sys.stderr)
    print(f"EVALUATION ({n} samples)", file=sys.stderr)
    print(f"  MAE:      {mae:.2f}", file=sys.stderr)
    print(f"  Pearson:  {corr:.3f}", file=sys.stderr)
    print(f"  Spearman: {sp:.3f}", file=sys.stderr)


if __name__ == "__main__":
    main()
